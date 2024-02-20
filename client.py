from datetime import datetime, date
from fastapi import FastAPI, Request
from argparse import ArgumentParser
from csv import DictReader
from json import dumps, loads
from requests import post
from openpyxl import Workbook, cell
from openpyxl.styles import Font, PatternFill


class RequestSender:
    def __init__(self, filepath):
        with open(filepath, newline="", encoding="utf-8") as csvdata:
            reader = DictReader(csvdata, delimiter=";")
            self.vehicle_data = [row for row in reader]

    def transmit_to_server(self):
        payload = dumps({"data": self.vehicle_data})
        headers = {"Content-Type": "application/json"}
        response = post("http://127.0.0.1:7000/", data=payload, headers=headers)
        response_dict = loads(response.json())
        return response_dict


class ExcelConventer:
    def __init__(self, resp_API, color_code, keys, colored) -> None:
        self.resp_API = resp_API
        self.keys = keys
        self.color_code = color_code
        self.colored = colored

    def adding_rows_excel(self, sheet, column_keys):
        curr_date = datetime.now()
        for index in range(2, len(self.resp_API) + 2):
            for key_val in range(0, len(column_keys)):
                cell = sheet.cell(
                    row=index,
                    column=key_val + 1,
                )
                sheet.cell(
                    row=index,
                    column=key_val + 1,
                    value=self.resp_API[index - 2][column_keys[key_val]],
                )
                color_cd = [
                    val["colorCode"]
                    for val in self.color_code
                    if val["labelIds"] == self.resp_API[index - 2]["labelIds"]
                ]
                if len(color_cd) > 0:
                    cell.font = Font(color=color_cd[0]["colorCode"])
                if self.colored == True:
                    date_item = datetime.strptime(
                        self.resp_API[index - 2]["hu"], "%Y-%m-%d"
                    )
                    diff = curr_date - date_item
                    if diff.days < 90:
                        cell.fill = PatternFill(
                            start_color="007500",
                            end_color="007500",
                            fill_type="solid",
                        )
                    elif diff.days < 12 * 30:
                        cell.fill = PatternFill(
                            start_color="FFA500",
                            end_color="FFA500",
                            fill_type="solid",
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="b30000",
                            end_color="b30000",
                            fill_type="solid",
                        )

    def convert_to_excel(self, filename):
        try:
            wb = Workbook()
            ws = wb.active
            column_keys = ["rnr"]
            if self.keys != None:
                for row_item in range(len(self.keys)):
                    if self.keys[row_item] not in self.resp_API[0].keys():
                        raise Exception("Key not exsist in dataset")
                    column_keys.append(self.keys[row_item])
            for row_item in range(0, len(column_keys)):
                ws.cell(row=1, column=row_item + 1, value=column_keys[row_item])
            self.adding_rows_excel(ws, column_keys)
            current_time = datetime.now()
            current_data_iso_formatted = current_time.isoformat()
            current_data_iso_formatted = current_data_iso_formatted.replace(":", "_")
            file_name = filename + "_{" + current_data_iso_formatted + "}.xlsx"
            wb.save(file_name)
        except Exception as error:
            print(error)


def main():
    try:
        ap = ArgumentParser(
            description="Transmit CSV to REST API and generate Excel file."
        )
        ap.add_argument("-k", "--keys", nargs="+", type=str)
        ap.add_argument("-c", "--colored", type=bool)
        parsed_arguments = ap.parse_args()
        file_path = "vehicles"  # The filepath can be changed from here.
        client = RequestSender(f"{file_path}.csv")
        response_data = client.transmit_to_server()
        print(response_data)
        excel_conventer = ExcelConventer(
            response_data["data"]["req_API"],
            response_data["data"]["color_code"],
            parsed_arguments.keys,
            parsed_arguments.colored,
        )
        excel_conventer.convert_to_excel(file_path)

    except Exception as e:
        print(e)


if __name__ == "__main__":
    main()
